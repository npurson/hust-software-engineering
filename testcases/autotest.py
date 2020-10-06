#!/usr/bin/python3
import argparse
import configparser
import csv
import json
import os
import psutil
import re
import signal
import subprocess
import time

from subprocess import Popen
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import tqdm
from tqdm.contrib.concurrent import process_map, thread_map

# basic config here
CONFIG_JSON_NAME = 'config.json'
CAPTURE_CMD = (
    'preset', 'user', 'map', 'fund', 'credit',
    'gift', 'bomb', 'barrier', 'userloc', 'nextuser'
)
DEFAULT_EXEC = '../richman/richman'
EXPIRE_TIME_MS = 5000
REPORT_FILE_NAME = 'report'
EOL_SEQ = '\n'
# end config

GREEN = PatternFill(start_color='7FFF00',
                    end_color='7FFF00', fill_type='solid')
RED = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

# arguments
args = argparse.ArgumentParser()
args.add_argument('-d', '--dir', help='testcase dir',
                  default='.', type=str)
args.add_argument('-w', '--workers',
                  help='numebr of work processes', default=4, type=int)
args.add_argument('-n', '--name', help='process name',
                  default=DEFAULT_EXEC, type=str)
args.add_argument('-e', '--export', help='export to file, 0=none, 1=csv, 2=xlsx',
                  default=2, type=int)
args.add_argument('-f', '--fail', help='export failed cases to file',
                  default=True, type=bool)
args.add_argument('-o', '--outputdir', help='output dir',
                  default='result', type=str)
args = args.parse_args()
# end arguments

base_dir = args.dir
exec_path = os.path.join(base_dir, args.name)
export_dir = os.path.join(base_dir, args.outputdir)

kv_pat = re.compile(r'([a-zA-Z]+)(?:\s+(\w+))(?:\s+(\w+))?(?:\s+(\w+))?')
test_list = dict()


def kill_proc_tree(p, timeout=None, on_terminate=None):
    '''
        p = pid of root process
    '''
    if isinstance(p, int):
        p = psutil.Process(p)
    elif not isinstance(p, psutil.Process):
        p = psutil.Process(p.pid)
    ch = [p]+p.children(recursive=True)
    for c in ch:
        c.kill()
    succ, err = psutil.wait_procs(ch, timeout=timeout, callback=on_terminate)
    return (succ, err)


def trim_line(line):
    if not line:
        return ''

    line = re.sub(r'#.*$', '', line)  # comment
    line = re.sub(r'\s+', ' ', line.strip())  # strip and merge empty char
    if not line:
        return ''
    line = line.lower()  # lowercase

    if not any(line.startswith(c + ' ') for c in CAPTURE_CMD):
        return ''  # check first word

    return line


def parse_key_value_pairs(lines):
    ret = []
    for line in lines:
        line = trim_line(line)
        if not line:
            continue
        mat = kv_pat.match(line)
        if mat:
            ret.append(' '.join((e for e in mat.groups() if e)))
        else:
            print('Warn::', line)
            # assert False
    # print(ret)
    ret.sort()
    return ret


def export_failed_case(filename, out_kv, ans_kv):
    i_out = 0
    i_ans = 0
    l_out = len(out_kv)
    l_ans = len(ans_kv)
    w_out = len('stderr')
    if l_out:
        w_out = max(max(len(e) for e in out_kv), w_out)

    fout = list()
    fout.append('stderr'.rjust(w_out, ' ') + '|expected')
        
    while i_out < l_out:
        if out_kv[i_out] == ans_kv[i_ans]:
            fout.append(out_kv[i_out].ljust(w_out, ' ') + '|' + ans_kv[i_ans])
            i_out = i_out + 1
            i_ans = i_ans + 1
        elif i_out < l_out:
            fout.append(out_kv[i_out].ljust(w_out, ' ') + '|')
            i_out = i_out + 1
        pass
    
    while i_ans < l_ans:
        fout.append(' ' * w_out + '|' + ans_kv[i_ans])
        i_ans = i_ans + 1
        pass

    with open(os.path.join(export_dir, filename), 'w') as f:
        f.write(EOL_SEQ.join(fout))
    return


def check(case):
    '''
        case[0] = test name
        case[1] = testcase path
    '''
    casename = os.path.basename(case[1])

    with open(os.path.join(base_dir, case[1]+'.in'), 'r') as f:
        testcase_input = f.read()
    with open(os.path.join(base_dir, case[1]+'.out'), 'r') as f:
        testcase_answer = f.read()

    proc = Popen(
        exec_path, shell=True, universal_newlines=True,
        stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE
    )

    output = ''
    try:
        t = time.time()
        _, output = proc.communicate(
            testcase_input + '\nquit\n',
            timeout=EXPIRE_TIME_MS/1000
        )
        t = time.time() - t
    except subprocess.TimeoutExpired:
        t = time.time() - t
        print('<timeout>', case[0], '@', case[1], ', pid', proc.pid)
        kill_proc_tree(proc.pid)

    # print('output', casename, '>', output)
    out_kv = parse_key_value_pairs(output.splitlines()) if output else list()
    ans_kv = parse_key_value_pairs(testcase_answer.splitlines())
    # print(ans_kv, out_kv)

    if args.fail and (not out_kv == ans_kv):
        # export error
        export_failed_case(
            'diff-%s-%s.txt' % (case[0], casename),
            out_kv, ans_kv
        )
        pass

    return (
        case[0],
        casename,
        out_kv == ans_kv,  # operator.eq(out_kv, ans_kv),
        t * 1e3
    )


def add_case(testname, casedir):
    print('<case>', testname, '@', casedir)
    if testname not in test_list.keys():
        test_list[testname] = list()
    test_list[testname].append(casedir)
    return


def get_all_case_filename(dir):
    return list(
        set(
            os.path.splitext(casename)[0]
            for casename in os.listdir(dir)
            if os.path.splitext(casename)[1] in ['.in', '.out']
        )
    )


def import_dir(currdir, basename='', maxlevel=1000):
    dir_list = os.listdir(currdir)
    if CONFIG_JSON_NAME in dir_list:
        # with json configuration file
        with open(os.path.join(currdir, CONFIG_JSON_NAME), 'r') as conf:
            conf = json.load(conf)

        name = os.path.basename(currdir)
        if 'name' in conf.keys():
            name = conf['name']

        if 'maxlevel' in conf.keys():
            maxlevel = min(maxlevel, int(conf['maxlevel']))

        if 'import' in conf.keys():
            # import other dir
            for i in conf['import']:
                d = os.path.join(currdir, i)
                if not os.path.isdir(d):
                    continue  # is not a dir
                n = ('%s.%s' % (basename, name)) if basename else name
                # print("import>", d, ',', n, ',', maxlevel)
                import_dir(d, n, maxlevel)
            pass

        if 'case' in conf.keys():
            if basename:
                n = '%s.%s' % (basename, name)
            else:
                n = name
            for item in conf['case']:
                if int(item[1]) <= maxlevel:
                    add_case(n, os.path.join(currdir, item[0]))

        pass
    elif maxlevel >= 0:
        # only testcases in dir
        c = get_all_case_filename(currdir)
        for item in c:
            add_case(
                '%s.%s' % (basename, os.path.basename(currdir)),
                os.path.join(currdir, item)
            )
    return


def export_to_csv(filename, res):
    with open(filename, 'w') as f:
        writer = csv.writer(f)
        res_keys = list(res.keys())
        res_keys.sort()
        for test in res_keys:
            cnt = 0
            for item in res[test]:
                if item[2]:
                    cnt = cnt + 1  # count True
            writer.writerow(['testname', test, '%d/%d' %
                             (cnt, len(res[test]))])
            writer.writerow(['testname', 'testcase', 'status', 'time(ms)'])
            writer.writerows(res[test])
            writer.writerow([])
        pass  # end with
    return


def export_to_xlsx(file_name, res):
    wb = Workbook()
    ws = wb.active  # work sheet

    row = 0
    res_keys = list(res.keys())
    res_keys.sort()
    for test in res_keys:
        cnt = 0
        for item in res[test]:
            if item[2]:
                cnt = cnt + 1  # count True

        row = row + 1
        ws.cell(row, 1).value = 'testname'
        ws.cell(row, 2).value = test
        ws.cell(row, 3).value = '%d/%d' % (cnt, len(res[test]))

        row = row + 1
        ws.cell(row, 1).value = 'testname'
        ws.cell(row, 2).value = 'testcase'
        ws.cell(row, 3).value = 'status'
        ws.cell(row, 4).value = 'time(ms)'

        for item in res[test]:
            row = row + 1
            ws.cell(row, 1).value = item[0]
            ws.cell(row, 2).value = item[1]
            ws.cell(row, 3).value = 'PASS' if item[2] else 'FAIL'
            ws.cell(row, 3).fill = GREEN if item[2] else RED
            ws.cell(row, 4).value = item[3]
            if item[3] >= EXPIRE_TIME_MS:
                ws.cell(row, 4).fill = RED
            pass  # end for

        row = row + 1  # empty line
        pass

    wb.save(file_name)
    return


def init_testcase(dir):
    import_dir(base_dir, '')
    # print(test_list)
    testcases = list()
    keys = list(test_list.keys())
    keys.sort()
    for k in keys:
        for v in test_list[k]:
            testcases.append((k, v))
    testcases = set(testcases)  # unique
    # print(testcases)
    return testcases


# start main
if not os.path.isdir(export_dir):
    os.mkdir(export_dir)

testcases = init_testcase(base_dir)

print('run', exec_path)
res = thread_map(check, testcases, max_workers=args.workers)

test_res = dict()
for item in res:
    if item[0] not in test_res:
        test_res[item[0]] = list()
    test_res[item[0]].append(item)
for item in test_res:
    test_res[item].sort(key=lambda x: x[1])

# report
print(test_res)
if args.export == 1:
    f = os.path.join(export_dir, REPORT_FILE_NAME + '.csv')
    print('export to', f)
    export_to_csv(f, test_res)
elif args.export == 2:
    f = os.path.join(export_dir, REPORT_FILE_NAME + '.xlsx')
    print('export to', f)
    export_to_xlsx(f, test_res)
